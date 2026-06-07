#!/usr/bin/env python3
"""Genereer QR-labels en een printbare PDF uit BootManager_LocationTags_Pilot.xlsx.

Gebruik:
    python generate-location-tag-prints.py
    python generate-location-tag-prints.py --excel BootManager_LocationTags_Pilot.xlsx --output output
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
import qrcode
from qrcode.constants import ERROR_CORRECT_Q
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from PIL import Image, ImageDraw, ImageFont


@dataclass(frozen=True)
class LocationTag:
    label_code: str
    guid: str
    qr_content: str
    location: str
    area: str
    placed: str
    note: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Genereer BootManager QR-labels en een printbare PDF."
    )
    parser.add_argument(
        "--excel",
        default="BootManager_LocationTags_Pilot.xlsx",
        help="Pad naar het Excel-bestand.",
    )
    parser.add_argument(
        "--output",
        default="generated-location-tags",
        help="Uitvoermap voor PNG-bestanden en PDF.",
    )
    parser.add_argument(
        "--sheet",
        default="Locatietags",
        help="Naam van het werkblad met de tags.",
    )
    parser.add_argument(
        "--include-empty",
        action="store_true",
        help="Neem ook regels zonder labelcode of QR-inhoud mee (standaard: overslaan).",
    )
    parser.add_argument(
        "--label-size-mm",
        type=float,
        default=40.0,
        help="Formaat van één vierkant label in millimeters (standaard: 40).",
    )
    return parser.parse_args()


def sanitize_filename(value: str) -> str:
    value = value.strip()
    value = re.sub(r"[^A-Za-z0-9._-]+", "_", value)
    return value or "tag"


def read_tags(excel_path: Path, sheet_name: str, include_empty: bool) -> list[LocationTag]:
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel-bestand niet gevonden: {excel_path}")

    workbook = load_workbook(excel_path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(
            f"Werkblad '{sheet_name}' niet gevonden. Beschikbaar: {', '.join(workbook.sheetnames)}"
        )

    sheet = workbook[sheet_name]
    headers = {
        str(cell.value).strip(): index
        for index, cell in enumerate(sheet[1], start=1)
        if cell.value is not None
    }

    required = ["Labelcode", "GUID", "QR-inhoud"]
    missing = [name for name in required if name not in headers]
    if missing:
        raise ValueError(f"Ontbrekende kolommen: {', '.join(missing)}")

    def get_value(row: int, header: str) -> str:
        col = headers.get(header)
        if col is None:
            return ""
        value = sheet.cell(row=row, column=col).value
        return "" if value is None else str(value).strip()

    tags: list[LocationTag] = []
    for row in range(2, sheet.max_row + 1):
        label_code = get_value(row, "Labelcode")
        qr_content = get_value(row, "QR-inhoud")

        if not include_empty and (not label_code or not qr_content):
            continue

        tags.append(
            LocationTag(
                label_code=label_code,
                guid=get_value(row, "GUID"),
                qr_content=qr_content,
                location=get_value(row, "Locatie"),
                area=get_value(row, "Gebied"),
                placed=get_value(row, "Geplaatst"),
                note=get_value(row, "Opmerking"),
            )
        )

    if not tags:
        raise ValueError("Geen bruikbare tags gevonden in het Excel-bestand.")

    duplicate_codes = find_duplicates(tag.label_code for tag in tags)
    duplicate_payloads = find_duplicates(tag.qr_content for tag in tags)
    if duplicate_codes:
        raise ValueError(f"Dubbele labelcodes gevonden: {', '.join(sorted(duplicate_codes))}")
    if duplicate_payloads:
        raise ValueError("Dubbele QR-inhoud gevonden; iedere QR moet uniek zijn.")

    return tags


def find_duplicates(values: Iterable[str]) -> set[str]:
    seen: set[str] = set()
    duplicates: set[str] = set()
    for value in values:
        if value in seen:
            duplicates.add(value)
        seen.add(value)
    return duplicates


def create_qr_image(payload: str, target_path: Path) -> None:
    qr = qrcode.QRCode(
        version=None,
        error_correction=ERROR_CORRECT_Q,
        box_size=12,
        border=4,
    )
    qr.add_data(payload)
    qr.make(fit=True)
    image = qr.make_image(fill_color="black", back_color="white")
    image.save(target_path)


def default_font(size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    candidates = [
        "DejaVuSans-Bold.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "C:/Windows/Fonts/arialbd.ttf",
    ]
    for candidate in candidates:
        try:
            return ImageFont.truetype(candidate, size=size)
        except OSError:
            continue
    return ImageFont.load_default()


def create_label_png(tag: LocationTag, qr_path: Path, target_path: Path) -> None:
    canvas_size = 1200
    margin = 70
    qr_size = 900
    label_height = canvas_size

    image = Image.new("RGB", (canvas_size, label_height), "white")
    draw = ImageDraw.Draw(image)
    qr_image = Image.open(qr_path).convert("RGB").resize((qr_size, qr_size))
    qr_x = (canvas_size - qr_size) // 2
    image.paste(qr_image, (qr_x, margin))

    code_font = default_font(58)
    detail_font = default_font(34)

    code_bbox = draw.textbbox((0, 0), tag.label_code, font=code_font)
    code_width = code_bbox[2] - code_bbox[0]
    draw.text(((canvas_size - code_width) / 2, 1000), tag.label_code, fill="black", font=code_font)

    detail_parts = [part for part in [tag.area, tag.location] if part]
    detail = " - ".join(detail_parts)
    if detail:
        if len(detail) > 42:
            detail = detail[:39] + "..."
        detail_bbox = draw.textbbox((0, 0), detail, font=detail_font)
        detail_width = detail_bbox[2] - detail_bbox[0]
        draw.text(
            ((canvas_size - detail_width) / 2, 1080),
            detail,
            fill="black",
            font=detail_font,
        )

    image.save(target_path, dpi=(300, 300))


def create_print_pdf(tags: list[LocationTag], label_dir: Path, pdf_path: Path, label_size_mm: float) -> None:
    page_width, page_height = A4
    margin = 12 * mm
    gap = 5 * mm
    label_size = label_size_mm * mm

    columns = max(1, int((page_width - 2 * margin + gap) // (label_size + gap)))
    rows = max(1, int((page_height - 2 * margin + gap) // (label_size + gap)))
    per_page = columns * rows

    pdf = canvas.Canvas(str(pdf_path), pagesize=A4)
    pdf.setTitle("BootManager locatietags")

    for index, tag in enumerate(tags):
        slot = index % per_page
        if index > 0 and slot == 0:
            pdf.showPage()

        row = slot // columns
        col = slot % columns

        x = margin + col * (label_size + gap)
        y = page_height - margin - (row + 1) * label_size - row * gap

        png_path = label_dir / f"{sanitize_filename(tag.label_code)}.png"
        pdf.drawImage(str(png_path), x, y, width=label_size, height=label_size, preserveAspectRatio=True)
        pdf.rect(x, y, label_size, label_size)

    pdf.save()


def main() -> int:
    args = parse_args()

    script_dir = Path(__file__).resolve().parent
    excel_path = Path(args.excel)
    if not excel_path.is_absolute():
        excel_path = script_dir / excel_path

    output_dir = Path(args.output)
    if not output_dir.is_absolute():
        output_dir = script_dir / output_dir

    qr_dir = output_dir / "qr-only"
    label_dir = output_dir / "labels"
    qr_dir.mkdir(parents=True, exist_ok=True)
    label_dir.mkdir(parents=True, exist_ok=True)

    try:
        tags = read_tags(excel_path, args.sheet, args.include_empty)

        for tag in tags:
            filename = sanitize_filename(tag.label_code)
            qr_path = qr_dir / f"{filename}.png"
            label_path = label_dir / f"{filename}.png"
            create_qr_image(tag.qr_content, qr_path)
            create_label_png(tag, qr_path, label_path)

        pdf_path = output_dir / "BootManager_LocationTags_Print.pdf"
        create_print_pdf(tags, label_dir, pdf_path, args.label_size_mm)

        print(f"Gereed: {len(tags)} tags gegenereerd.")
        print(f"QR-bestanden: {qr_dir}")
        print(f"Labelafbeeldingen: {label_dir}")
        print(f"Printbare PDF: {pdf_path}")
        return 0
    except Exception as exc:  # noqa: BLE001
        print(f"FOUT: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
